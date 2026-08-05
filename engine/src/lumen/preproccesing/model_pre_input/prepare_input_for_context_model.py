# #[AI_CONTEXT_START]
# - CONFIGURACIÓN DE FACTORY: `ModelPreInputFactory` y contenedor DI del módulo, inyectados por la Factory Maestra de Agentes.
# - ABSTRACCIÓN DEL DATO: Schema de entrada robusto tipado por backend, no dicts ad hoc.
# - REFACTOR NATIVO: Tests del pipeline de contexto ML en los tres backends configurables.
# #[AI_CONTEXT_END]
import os
import csv
import json
import logging
from abc import ABC, abstractmethod
from dataclasses import dataclass
from typing import List, Optional
from urllib.parse import urlparse
from langchain_core.tools import tool
from pydantic import BaseModel, Field

# ==========================================
# 0. INPUT SCHEMA PARA LANGCHAIN (Agregado)
# ==========================================
#TODO [AI] : CHANGE THIS FOR MORE ROBUZT SCHEMA 
class MetadataToolInputSchema(BaseModel):
    """Esquema Pydantic obligatorio para que LangChain valide el input del agente."""
    path: str = Field(
        ..., 
        description="Ruta absoluta, relativa o URI del archivo de datos a analizar."
    )

# ==========================================
# 1. TOOL DE LANGCHAIN (Adaptada al schema)
# ==========================================
@tool(args_schema=MetadataToolInputSchema, return_direct=False)
def meta_data_context(path: str) -> str:
    """Extrae metadatos de la fuente de datos."""
    # Instanciamos usando tu contenedor de dependencias original
    data_context_service = DIContainer.build_data_context_service()
    # Usamos tu método generate_agent_prompt
    return data_context_service.generate_agent_prompt(
        uri_or_path=path, 
        base_system_prompt="Please analyze the following source metadata:"
    )

# ==========================================
# 2. DOMAIN MODELS & INTERFACES (Unchanged)
# ==========================================

@dataclass
class SourceMetadata:
    source_type: str
    recommended_backend: str
    detected_columns: List[str]
    estimated_size: str
    is_structured: bool
    schema_details: str = ""

    def to_prompt_format(self) -> str:
        cols = f"{self.detected_columns[:10]}... (Truncated)" if len(self.detected_columns) > 10 else self.detected_columns
        return (
            f"--- RAW SOURCE METADATA ---\n"
            f"Source Format: {self.source_type}\n"
            f"Structure Status: {'Structured/Tabular' if self.is_structured else 'Semi/Unstructured'}\n"
            f"Estimated Size: {self.estimated_size}\n"
            f"Detected Schema/Columns: {cols}\n"
            f"Details: {self.schema_details}\n"
            f"-> RECOMMENDED TOOL BACKEND: {self.recommended_backend}\n"
            f"---------------------------"
        )

class ISourceInspector(ABC):
    @abstractmethod
    def can_handle(self, uri_or_path: str) -> bool:
        pass

    @abstractmethod
    def inspect(self, uri_or_path: str) -> SourceMetadata:
        pass

class IInspectorFactory(ABC):
    @abstractmethod
    def get_inspector(self, uri_or_path: str) -> ISourceInspector:
        pass

# ==========================================
# 3. EXPANDED CONCRETE STRATEGIES (Inspectors)
# ==========================================

class CSVInspector(ISourceInspector):
    def can_handle(self, uri_or_path: str) -> bool:
        return uri_or_path.lower().endswith(('.csv', '.tsv'))

    def inspect(self, uri_or_path: str) -> SourceMetadata:
        size_mb = os.path.getsize(uri_or_path) / (1024 * 1024) if os.path.exists(uri_or_path) else 0.0
        headers = []
        if os.path.exists(uri_or_path):
            with open(uri_or_path, 'r', encoding='utf-8') as file:
                headers = next(csv.reader(file), [])
        
        backend = "AnalyzeContextSpark" if size_mb > 1000 else "AnalyzeContextPolars"
        return SourceMetadata("CSV Flat File", backend, headers, f"{size_mb:.2f} MB", True)

class JSONInspector(ISourceInspector):
    def can_handle(self, uri_or_path: str) -> bool:
        return uri_or_path.lower().endswith(('.json', '.jsonl'))

    def inspect(self, uri_or_path: str) -> SourceMetadata:
        size_mb = os.path.getsize(uri_or_path) / (1024 * 1024) if os.path.exists(uri_or_path) else 0.0
        keys = []
        if os.path.exists(uri_or_path):
            with open(uri_or_path, 'r', encoding='utf-8') as file:
                try:
                    data = json.load(file)
                    if isinstance(data, dict): keys = list(data.keys())
                    elif isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict): keys = list(data[0].keys())
                except json.JSONDecodeError:
                    pass # Handle empty or malformed gracefully
        return SourceMetadata("JSON Document", "AnalyzeContextPandas", keys, f"{size_mb:.2f} MB", False, "Nested key-value structure.")

class ParquetInspector(ISourceInspector):
    """Inspects Columnar Big Data formats by reading the lightweight metadata footer."""
    def can_handle(self, uri_or_path: str) -> bool:
        return uri_or_path.lower().endswith('.parquet')

    def inspect(self, uri_or_path: str) -> SourceMetadata:
        try:
            import pyarrow.parquet as pq
            size_mb = os.path.getsize(uri_or_path) / (1024 * 1024) if os.path.exists(uri_or_path) else 0.0
            
            columns = []
            if os.path.exists(uri_or_path):
                schema = pq.read_schema(uri_or_path)
                columns = schema.names

            backend = "AnalyzeContextSpark" if size_mb > 1000 else "AnalyzeContextPolars"
            return SourceMetadata("Parquet Binary (Columnar)", backend, columns, f"{size_mb:.2f} MB", True, f"{len(columns)} columns mapped.")
        except ImportError:
            raise ImportError("Critical: 'pyarrow' library is required to inspect Parquet files.")

class ExcelInspector(ISourceInspector):
    """Inspects Spreadsheets without loading all sheets into memory."""
    def can_handle(self, uri_or_path: str) -> bool:
        return uri_or_path.lower().endswith(('.xlsx', '.xls', '.ods'))

    def inspect(self, uri_or_path: str) -> SourceMetadata:
        try:
            import openpyxl
            size_mb = os.path.getsize(uri_or_path) / (1024 * 1024) if os.path.exists(uri_or_path) else 0.0
            
            headers = []
            sheet_names = []
            if os.path.exists(uri_or_path):
                # read_only=True prevents memory bloat
                wb = openpyxl.load_workbook(uri_or_path, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                if sheet_names:
                    sheet = wb[sheet_names[0]]
                    headers = [str(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                wb.close()

            return SourceMetadata("Excel Spreadsheet", "AnalyzeContextPandas", headers, f"{size_mb:.2f} MB", True, f"Detected sheets: {sheet_names}")
        except ImportError:
            raise ImportError("Critical: 'openpyxl' is required to inspect Excel files.")

class XMLInspector(ISourceInspector):
    """Inspects XML structures using the standard library."""
    def can_handle(self, uri_or_path: str) -> bool:
        return uri_or_path.lower().endswith('.xml')

    def inspect(self, uri_or_path: str) -> SourceMetadata:
        import xml.etree.ElementTree as ET
        size_mb = os.path.getsize(uri_or_path) / (1024 * 1024) if os.path.exists(uri_or_path) else 0.0
        
        tags = set()
        if os.path.exists(uri_or_path):
            # Parse iteratively to avoid OOM on massive XMLs
            context = ET.iterparse(uri_or_path, events=('start',))
            for i, (_, elem) in enumerate(context):
                tags.add(elem.tag)
                elem.clear()
                if i > 50: break # Only sample the first 50 elements

        return SourceMetadata("XML Document", "AnalyzeContextPandas", list(tags), f"{size_mb:.2f} MB", False, "Hierarchical tree structure.")

class SQLDatabaseInspector(ISourceInspector):
    """Inspects relational DB catalogs (PostgreSQL, MySQL, SQLite, SQL Server)."""
    def can_handle(self, uri_or_path: str) -> bool:
        parsed = urlparse(uri_or_path)
        return parsed.scheme in ['sqlite', 'postgresql', 'mysql', 'mssql', 'oracle']

    def inspect(self, uri_or_path: str) -> SourceMetadata:
        try:
            from sqlalchemy import create_engine, inspect
            
            engine = create_engine(uri_or_path)
            inspector = inspect(engine)
            tables = inspector.get_table_names()
            
            columns = []
            if tables:
                columns = [col['name'] for col in inspector.get_columns(tables[0])]

            return SourceMetadata(
                source_type="Relational Database (SQL)",
                recommended_backend="AnalyzeContextPandas",
                detected_columns=columns,
                estimated_size="Unknown (Remote DB)",
                is_structured=True,
                schema_details=f"Tables detected: {tables}. Columns sampled from '{tables[0]}'." if tables else "Empty Database."
            )
        except ImportError:
            raise ImportError("Critical: 'sqlalchemy' and specific DB drivers (like psycopg2) are required.")

class MongoNoSQLInspector(ISourceInspector):
    """Inspects Document-based NoSQL Databases."""
    def can_handle(self, uri_or_path: str) -> bool:
        return uri_or_path.startswith('mongodb://') or uri_or_path.startswith('mongodb+srv://')

    def inspect(self, uri_or_path: str) -> SourceMetadata:
        try:
            import pymongo
            
            client = pymongo.MongoClient(uri_or_path, serverSelectionTimeoutMS=2000)
            db = client.get_default_database() # Requires DB name in URI
            collections = db.list_collection_names()
            
            keys = set()
            if collections:
                sample_doc = db[collections[0]].find_one()
                if sample_doc:
                    keys = set(sample_doc.keys())

            return SourceMetadata(
                source_type="NoSQL Database (MongoDB)",
                recommended_backend="AnalyzeContextPandas",
                detected_columns=list(keys),
                estimated_size="Unknown (Remote DB)",
                is_structured=False,
                schema_details=f"Collections: {collections}. Sampled keys from '{collections[0]}'." if collections else "Empty Database."
            )
        except ImportError:
            raise ImportError("Critical: 'pymongo' is required to inspect MongoDB URIs.")

class CloudStorageInspector(ISourceInspector):
    """Handles Data Lake Object URIs."""
    def can_handle(self, uri_or_path: str) -> bool:
        parsed = urlparse(uri_or_path)
        return parsed.scheme in ['s3', 'gs', 'azure']

    def inspect(self, uri_or_path: str) -> SourceMetadata:
        # In a real enterprise app, you'd use boto3 (AWS) or google-cloud-storage here
        # to fetch the object metadata header. For now, we infer from the URI extension.
        inferred_type = uri_or_path.split('.')[-1] if '.' in uri_or_path else "Object File"
        
        backend = "AnalyzeContextSpark" # Cloud data lakes generally imply Spark/Big Data workflows
        if inferred_type.lower() in ['csv', 'json']:
            backend = "AnalyzeContextPandas"
            
        return SourceMetadata(
            source_type=f"Cloud Storage Object ({urlparse(uri_or_path).scheme.upper()})",
            recommended_backend=backend,
            detected_columns=["(Requires Cloud Authentication to Read Schema)"],
            estimated_size="Unknown (Remote Object)",
            is_structured=inferred_type.lower() in ['csv', 'parquet', 'orc'],
            schema_details=f"Target file extension inferred as: .{inferred_type}"
        )

class FallbackInspector(ISourceInspector):
    """Catch-all for unknown formats. Must be registered last."""
    def can_handle(self, uri_or_path: str) -> bool:
        return True

    def inspect(self, uri_or_path: str) -> SourceMetadata:
        return SourceMetadata("Unknown/Unsupported Format", "AnalyzeContextPandas", [], "Unknown", False, "Fallback protocol activated.")

# ==========================================
# 4. FACTORY AND SERVICE LAYER
# ==========================================

class SourceInspectorFactory(IInspectorFactory):
    def __init__(self) -> None:
        self._inspectors: List[ISourceInspector] = []

    def register_inspector(self, inspector: ISourceInspector) -> None:
        self._inspectors.append(inspector)

    def get_inspector(self, uri_or_path: str) -> ISourceInspector:
        for inspector in self._inspectors:
            if inspector.can_handle(uri_or_path):
                # Usando un print fallback en lugar de logger si logger no está inicializado en este script
                # logger.info(f"Factory routed '{uri_or_path}' to {type(inspector).__name__}")
                return inspector
        raise ValueError(f"Factory Configuration Error: No registered inspector can handle '{uri_or_path}'")

class DataContextService:
    def __init__(self, factory: IInspectorFactory) -> None:
        self._factory = factory

    def generate_agent_prompt(self, uri_or_path: str, base_system_prompt: str) -> str:
        inspector = self._factory.get_inspector(uri_or_path)
        metadata = inspector.inspect(uri_or_path)
        return f"{base_system_prompt}\n\n{metadata.to_prompt_format()}"

# ==========================================
# 5. DEPENDENCY INJECTION CONTAINER
# ==========================================

class DIContainer:
    """The Composition Root mapping all supported formats to the factory."""
    @staticmethod
    def build_data_context_service() -> DataContextService:
        factory = SourceInspectorFactory()
        
        # Register in order of specificity. 
        # CAUTION: FallbackInspector MUST be registered last.
        factory.register_inspector(CSVInspector())
        factory.register_inspector(JSONInspector())
        factory.register_inspector(ParquetInspector())
        factory.register_inspector(ExcelInspector())
        factory.register_inspector(XMLInspector())
        factory.register_inspector(SQLDatabaseInspector())
        factory.register_inspector(MongoNoSQLInspector())
        factory.register_inspector(CloudStorageInspector())
        
        # Always register the Fallback last so it only catches if all else fails
        factory.register_inspector(FallbackInspector())
        
        return DataContextService(factory=factory)